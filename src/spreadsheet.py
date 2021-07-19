from decimal import Decimal
from typing import Callable, Sequence
from openpyxl import Workbook

from openpyxl.worksheet.worksheet import Worksheet
from compile import Finances

from datainput import Transaction
from util import USD, percent

default_n = 5


def tab_append_top_trans(finances: Finances, tab: Worksheet, data: Sequence[int], top_n=default_n):
    return tab_append_sort_key(finances, tab, data, top_n, lambda i: abs(finances.ix(i).amt))


def tab_append_sort_key(finances: Finances, tab: Worksheet, data: Sequence[int], top_n: int, sort_fn: Callable[[int], Decimal]):
    sorted_data = sorted(data, key=sort_fn, reverse=True)
    return tab_append_data(finances, tab, sorted_data[:top_n])


def tab_append_data(finances: Finances, tab: Worksheet, data: Sequence[int]):
    trans = map(lambda i: finances.ix(i), data)
    tab_append_trans(tab, trans)


def tab_append_trans(tab, data: Sequence[Transaction]):
    tab.append(["Date", "Amount", "Class", "Description", "Source"])
    for t in sorted(data, key=lambda t: abs(t.amt), reverse=True):
        tab.append([t.date, USD(t.amt), t.classification, t.desc, t.source])


def create_spreadsheet(finances: Finances, output_filepath: str):
    wb = Workbook()
    summaryws = wb.active

    # Create year summary tab
    period = finances.whole_finances
    summaryws.title = "Period Summary"
    summaryws.append(["Since", f"{str(period.cutoff_date)}"])
    summaryws.append([])

    summaryws.append(["Income", f"{USD(period.income.income_sum)}"])
    summaryws.append(["Work Income", f"{USD(period.income.workincomesum)}"])
    summaryws.append(
        ["Non-Work Income", f"{USD(period.income.nonworkincomesum)}"])
    tab_append_top_trans(finances, summaryws, period.income.nonworkincome)
    summaryws.append([])

    summaryws.append(
        ["Returns Income", f"{USD(period.income.returnsincomesum)}"])
    summaryws.append(["Expenses", f"{USD(period.expense.expensessum)}"])
    summaryws.append([])

    summaryws.append(
        ["Gross Savings", f"{USD(period.gross_savings.amount_saved)}"])
    summaryws.append(["Gross Savings Percentage",
                     f"{percent(period.gross_savings.saving_ratio)}"])
    summaryws.append(["Gross Avg Monthly Savings",
                     f"{USD(period.avg_monthly_savings.amount_saved)}"])
    summaryws.append(
        [f"Gross Avg Monthly Savings Goals for {period.period_in_months} months"])
    summaryws.append(["Goal", "Goal Amount", "Leftover"])
    sav = finances.whole_finances.avg_monthly_savings
    for i in range(len(sav.savings_goals)):
        summaryws.append([percent(sav.savings_goals[i]), USD(
            sav.savings_goal_amts[i]), USD(sav.savings_goal_leftover[i])])
    summaryws.append([])

    summaryws.append(["Investments", f"{USD(period.investments_sum)}"])
    summaryws.append([])

    summaryws.append(["Category", f"% of Expenses", "Amount"])
    egc = period.expense.expensegroupcoll
    for eg_i in egc.group_sums_sorted[:20]:
        summaryws.append([egc.group_keys[eg_i], percent(
            egc.group_percentages[eg_i]), USD(egc.group_sums[eg_i])])
    summaryws.append([])

    # Period Income tab
    tab = wb.create_sheet()
    tab.title = "Period Income"
    tab_append_data(finances, tab, period.income.income)

    # Period Work Income tab
    tab = wb.create_sheet()
    tab.title = "Period Work Income"
    tab_append_data(finances, tab, period.income.workincome)

    # Period Non-work Income tab
    tab = wb.create_sheet()
    tab.title = "Period Non-work Income"
    tab_append_data(finances, tab, period.income.nonworkincome)

    # Period Returns Income tab
    tab = wb.create_sheet()
    tab.title = "Period Returns Income"
    tab_append_data(finances, tab, period.income.returnsincome)

    # Period Expenses tab
    tab = wb.create_sheet()
    tab.title = "Period Expenses"
    tab_append_data(finances, tab, period.expense.expenses)

    # Period Investments tab
    tab = wb.create_sheet()
    tab.title = "Period Investments"
    tab_append_data(finances, tab, period.investments)

    # Monthly Tabs
    sorted_month_keys = sorted(finances.monthly_finances.keys(), reverse=True)
    for month_key in sorted_month_keys:
        month = finances.monthly_finances[month_key]
        # Month Summary
        tab = wb.create_sheet()
        tab.title = f"{str(month_key)} Summary"
        tab.append(["Income", USD(month.income.income_sum)])
        tab.append(["Work Income", USD(month.income.workincomesum)])
        tab_append_top_trans(finances, tab, month.income.workincome)
        tab.append([])
        tab.append(["Returns Income", USD(month.income.returnsincomesum)])
        tab.append(["Expenses", USD(month.expense.expensessum)])
        tab.append(["Gross Savings", USD(month.savings.amount_saved)])
        tab.append(["Gross Savings Percentage",
                   percent(month.savings.saving_ratio)])
        tab.append(["Top expenses"])
        tab_append_top_trans(finances, tab, month.expense.expenses, 10)
        tab.append([])
        tab.append(["Top expense groups"])
        egc = month.expense.expensegroupcoll
        for eg_i in egc.group_sums_sorted[:5]:
            tab.append([egc.group_keys[eg_i], USD(egc.group_sums[eg_i])])

        # Month Data
        tab = wb.create_sheet()
        tab.title = f"{str(month_key)} Data"
        tab_append_data(finances, tab, month.source)

    # Ignored Transactions tab
    tab = wb.create_sheet()
    tab.title = "Ignored Transactions"
    tab_append_data(finances, tab, finances.ignored)

    # Source Data tab
    tab = wb.create_sheet()
    tab.title = "Period Source Data"
    tab_append_trans(tab, finances.source)

    wb.save(output_filepath)
