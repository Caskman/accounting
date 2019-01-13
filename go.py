from openpyxl import Workbook
from datetime import datetime
from operator import attrgetter, itemgetter

from datainput import get_data
from datafilter import filter_data

def main():
    data = get_data()

    wb = Workbook()
    summaryws = wb.active
    incomews = wb.create_sheet()
    expensews = wb.create_sheet()

    filtered_data = filter_data(data)
    leftover_data = set(data) - set(filtered_data)
    data = filtered_data

    income_data = [d for d in data if d.amt > 0]
    expense_data = [d for d in data if d.amt < 0]

    incomews.title = "Income"
    incomews.append(["Date", "Description", "Amount"])
    for d in sorted(income_data, key=attrgetter("date"), reverse=True):
        incomews.append([d.date, d.desc, d.amt])
    incomews.append(["", "", f"=SUM(C2:C{incomews.max_row})"])
    income_sum_addr = f"{incomews.title}!C{incomews.max_row}"

    expensews.title = "Expenses"
    expensews.append(["Date", "Description", "Amount"])
    for d in sorted(expense_data, key=attrgetter("date"), reverse=True):
        expensews.append([d.date, d.desc, d.amt])
    expensews.append(["", "", f"=SUM(C2:C{expensews.max_row})"])
    expense_sum_addr = f"{expensews.title}!C{expensews.max_row}"

    summaryws.title = "Summary"
    summaryws.append(["Income", f"={income_sum_addr}"])
    summaryws.append(["Expense", f"={expense_sum_addr}"])
    summaryws.append(["Net", f"=B1+B2"])
    summaryws.append(["Savings Rate", f"=100*B3/B1"])

    months_dict = {}
    def months_key(date_obj):
        return datetime.strptime(f"{date_obj.year}-{date_obj.month}", "%Y-%m").date()
    for d in data:
        key = months_key(d.date)
        month_obj = None
        if key in months_dict:
            month_obj = months_dict[key]
        else:
            month_obj = Month(key, [],[])
            months_dict[key] = month_obj
        if d.amt > 0:
            month_obj.income.append(d)
        else:
            month_obj.expense.append(d)

    summaryws.append([])
    summaryws.append(["Date", "Income", "Expense", "Net"])
    months_list = [(k, months_dict[k]) for k in months_dict]
    for month_t in sorted(months_list, key=itemgetter(0), reverse=True):
        date = month_t[0]
        month = month_t[1]
        income = sum(map(lambda x: x.amt, month.income))
        expense = sum(map(lambda x: x.amt, month.expense))
        row = summaryws.max_row + 1
        summaryws.append([date, income, expense, f"=B{row}+C{row}"])

        ws = wb.create_sheet()
        ws.title = datetime.strftime(date, "%Y-%m")
        ws.append(["Income"])
        ws.append(["Date", "Description", "Amount"])
        for item in sorted(month.income, key=attrgetter("amt"), reverse=True):
            ws.append([item.date, item.desc, item.amt])
        ws.append(["", "", f"=SUM(C3:C{ws.max_row})"])
        income_sum_addr = f"C{ws.max_row}"

        ws.append([])
        ws.append(["Expenses"])
        ws.append(["Date", "Description", "Amount", "Income %"])
        for item in sorted(month.expense, key=attrgetter("amt")):
            ws.append([item.date, item.desc, item.amt, f"=-100*C{ws.max_row + 1}/{income_sum_addr}"])
    
    leftoverws = wb.create_sheet()
    leftoverws.title = "Leftover Data"

    leftoverws.append(["Date", "Description", "Amount"])
    for item in sorted(leftover_data, key=attrgetter("date"), reverse=True):
        leftoverws.append([item.date, item.desc, item.amt])


    wb.save("output.xlsx")



class Month():
    def __init__(self, date, income, expense):
        self.date = date
        self.income = income
        self.expense = expense

if __name__ == "__main__":
    main()

